import re
import sys
import os
import time
import serial
import threading
from datetime import datetime

import matplotlib
matplotlib.use("Qt5Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

from PyQt5.QtCore import Qt, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QPlainTextEdit
)

################################################################################
# ExcelRecorder (unchanged from before, adapted for your new line format)
################################################################################
class ExcelRecorder:
    regex_time_s  = re.compile(r"([\d\.]+)\s*s")   
    regex_pwm     = re.compile(r"PWM\s*:\s*([\d]+)\s*/\s*([\d]+)")
    regex_t1      = re.compile(r"t1:\s*([\d\.]+)")
    regex_t2      = re.compile(r"t2:\s*([\d\.]+)")
    regex_t3      = re.compile(r"t3:\s*([\d\.]+)")
    regex_t3_est  = re.compile(r"t3\s*est:\s*([\d\.]+)")
    regex_t4      = re.compile(r"t4:\s*([\d\.]+)")

    def __init__(self, template_path, sheet_name="data"):
        self.template_path = template_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.current_path = None
        self.row_index = 2
        self.base_time = None
        self.t3_values = []
        self.consigne = 25.0

    def create_copy_and_open(self, new_path):
        import shutil
        from openpyxl import load_workbook

        shutil.copyfile(self.template_path, new_path)
        self.current_path = new_path

        self.workbook = load_workbook(new_path)
        self.sheet = self.workbook[self.sheet_name]
        self.row_index = 2
        self.base_time = None
        self.t3_values.clear()

    def set_consigne(self, value):
        self.consigne = float(value)

    def parse_and_write(self, line):
        m_s = self.regex_time_s.search(line)
        print(line)
        if not m_s:
            return
        time_s = float(m_s.group(1))

        pwm_m = self.regex_pwm.search(line)
        if not pwm_m:
            return
        duty_val = float(pwm_m.group(1))
        duty_max = float(pwm_m.group(2))
        fraction = duty_val / duty_max if duty_max else 0.0
        U = fraction * 10.0 - 5.0

        mt1 = self.regex_t1.search(line)
        mt2 = self.regex_t2.search(line)
        mt3 = self.regex_t3.search(line)
        mt4 = self.regex_t4.search(line)
        if not (mt1 and mt2 and mt3 and mt4):
            return

        t1_val = float(mt1.group(1))
        t2_val = float(mt2.group(1))
        t3_val = float(mt3.group(1))
        t4_val = float(mt4.group(1))

        mt3e = self.regex_t3_est.search(line)
        t3_est_val = float(mt3e.group(1)) if mt3e else None

        if self.base_time is None:
            self.base_time = time_s

        temps = time_s - self.base_time
        self.t3_values.append(t3_val)
        t3_moy = sum(self.t3_values) / len(self.t3_values)

        # Write row
        row = self.row_index
        self.sheet.cell(row=row, column=1).value = temps
        self.sheet.cell(row=row, column=2).value = self.consigne
        self.sheet.cell(row=row, column=3).value = U
        self.sheet.cell(row=row, column=4).value = t1_val
        self.sheet.cell(row=row, column=5).value = t2_val
        self.sheet.cell(row=row, column=6).value = t3_val
        self.sheet.cell(row=row, column=7).value = t3_est_val if t3_est_val else ""
        self.sheet.cell(row=row, column=8).value = t4_val
        self.sheet.cell(row=row, column=9).value = t3_moy

        self.row_index += 1

    def save_and_close(self):
        if self.workbook:
            self.workbook.save(self.current_path)
            self.workbook = None
            self.sheet = None


################################################################################
# Thread for reading data from the serial port
################################################################################
class SerialReadThread(threading.Thread):
    def __init__(self, ser, callback):
        super().__init__()
        self.ser = ser
        self.callback = callback
        self.running = True

    def run(self):
        while self.running:
            try:
                line = self.ser.readline().decode('ascii', errors='ignore').strip()
                if line:
                    self.callback(line)
            except Exception as e:
                print("Serial error:", e)
                break
        print("Serial reading thread terminated.")

    def stop(self):
        self.running = False


################################################################################
# Live Plot #1: Temperatures (T1..T4 + T3_est) [no stability text here]
################################################################################
class TempPlotWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        # Data arrays
        self.time_values     = []
        self.consigne_values = []
        self.t1_values       = []
        self.t2_values       = []
        self.t3_values       = []
        self.t4_values       = []
        self.t3_est_values   = []

        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)

        self.line_consigne,     = self.ax.plot([], [], label="Consigne")
        self.line_t1,     = self.ax.plot([], [], label="T1")
        self.line_t2,     = self.ax.plot([], [], label="T2")
        self.line_t3,     = self.ax.plot([], [], label="T3")
        self.line_t4,     = self.ax.plot([], [], label="T4")
        self.line_t3_est, = self.ax.plot([], [], label="T3_est")

        self.ax.set_xlabel("Time (s)")
        self.ax.set_ylabel("Temperature (°C)")
        self.ax.set_title("Real-Time Temps + T3_est")
        self.ax.grid(True)
        self.ax.legend()

        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.setLayout(layout)

    @pyqtSlot(str)
    def parse_and_update(self, line):
        match_s = re.search(r"([\d\.]+)\s*s", line)
        if not match_s:
            return
        time_s = float(match_s.group(1))
   

        # t1..t4
        mt1 = re.search(r"t1:\s*([\d\.]+)", line)
        mt2 = re.search(r"t2:\s*([\d\.]+)", line)
        mt3 = re.search(r"t3:\s*([\d\.]+)", line)
        mt4 = re.search(r"t4:\s*([\d\.]+)", line)
        if not (mt1 and mt2 and mt3 and mt4):
            return
        
        t1_val = float(mt1.group(1))
        t2_val = float(mt2.group(1))
        t3_val = float(mt3.group(1))
        t4_val = float(mt4.group(1))

        # t3 est (optional)
        mt3e = re.search(r"t3\s*est:\s*([\d\.]+)", line)
        t3_est_val = float(mt3e.group(1)) if mt3e else None

        consigne = re.search(r"consigne:\s*([\d\.]+)", line)
        consigne_val = float(consigne.group(1)) if consigne else None
        
        # Append
        self.time_values.append(time_s)
        self.t1_values.append(t1_val)
        self.t2_values.append(t2_val)
        self.t3_values.append(t3_val)
        self.t4_values.append(t4_val)

        if t3_est_val is not None:
            self.t3_est_values.append(t3_est_val)
        else:
            self.t3_est_values.append(None)
            
        if consigne_val is not None:
            self.consigne_values.append(consigne_val)
        else:
            self.consigne_values.append(None)

        # Update lines
        self.line_consigne.set_xdata(self.time_values)
        self.line_consigne.set_ydata([val if val is not None else float("nan") for val in self.consigne_values])
        # self.line_consigne.set_ydata(self.consigne_values)
        self.line_t1.set_xdata(self.time_values)
        self.line_t1.set_ydata(self.t1_values)
        self.line_t2.set_xdata(self.time_values)
        self.line_t2.set_ydata(self.t2_values)
        self.line_t3.set_xdata(self.time_values)
        self.line_t3.set_ydata(self.t3_values)
        # self.line_t4.set_xdata(self.time_values)
        # self.line_t4.set_ydata(self.t4_values)

        # T3_est -> None => NaN
        t3_est_y = [val if val is not None else float("nan") for val in self.t3_est_values]
        self.line_t3_est.set_xdata(self.time_values)
        self.line_t3_est.set_ydata(t3_est_y)

        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()

    def reset_plot(self):
        self.time_values.clear()
        self.consigne_values.clear()
        self.t1_values.clear()
        self.t2_values.clear()
        self.t3_values.clear()
        self.t4_values.clear()
        self.t3_est_values.clear()

        self.line_consigne.set_data([], [])
        self.line_t1.set_data([], [])
        self.line_t2.set_data([], [])
        self.line_t3.set_data([], [])
        self.line_t4.set_data([], [])
        self.line_t3_est.set_data([], [])

        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()


################################################################################
# Live Plot #2: Command (U)
################################################################################
class CommandPlotWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.time_values = []
        self.u_values    = []

        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)

        self.line_u, = self.ax.plot([], [], label="U (%)")

        self.ax.set_xlabel("Time (s)")
        self.ax.set_ylabel("Command U (%)")
        self.ax.set_title("Real-Time Command U")
        self.ax.grid(True)
        self.ax.legend()

        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.setLayout(layout)

    @pyqtSlot(str)
    def parse_and_update(self, line):
        match_s = re.search(r"([\d\.]+)\s*s", line)
        if not match_s:
            return
        time_s = float(match_s.group(1))


        pwm_m = re.search(r"PWM\s*:\s*([\d]+)\s*/\s*([\d]+)", line)
        if not pwm_m:
            return
        duty_val = float(pwm_m.group(1))
        duty_max = float(pwm_m.group(2))
        ratio = duty_val/duty_max if duty_max else 0.0
        u_val = -ratio * 200 + 100

        self.time_values.append(time_s)
        self.u_values.append(u_val)

        self.line_u.set_xdata(self.time_values)
        self.line_u.set_ydata(self.u_values)

        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()

    def reset_plot(self):
        self.time_values.clear()
        self.u_values.clear()

        self.line_u.set_data([], [])
        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()


################################################################################
# Main PyQt Window
################################################################################
class SerialMonitor(QWidget):
    lineReceivedText = pyqtSignal(str)
    lineReceivedPlot = pyqtSignal(str)

    def __init__(self, port="COM4", baudrate=115200, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Design 2 Prototype serial monitor")
        self.setWindowIcon(QIcon("icon.png"))

        self.data_dir = os.path.join(os.getcwd(), "data")
        os.makedirs(self.data_dir, exist_ok=True)

        self.recording = False
        self.file = None

        self.template_file = os.path.join(os.getcwd(), "app/assets/data.xlsx")
        self.excel_recorder = ExcelRecorder(self.template_file)

        # For stability tracking
        self.current_setpoint = 25.0
        self.allowable_error  = 1.25
        self.in_stable_zone   = False
        self.enter_time       = None

        self._build_ui()

        self.lineReceivedText.connect(self.append_line)
        self.lineReceivedPlot.connect(self.temp_plot.parse_and_update)
        self.lineReceivedPlot.connect(self.cmd_plot.parse_and_update)

        # Try open serial
        try:
            self.ser = serial.Serial(port, baudrate, timeout=1)
            time.sleep(2)
        except Exception as e:
            raise RuntimeError(f"Could not open port {port}: {e}")

        self.read_thread = SerialReadThread(self.ser, self.on_line_received)
        self.read_thread.start()

    def _build_ui(self):
        main_layout = QHBoxLayout()
        control_layout = QVBoxLayout()
        plot_layout = QVBoxLayout()

        # Buttons
        btn_layout = QHBoxLayout()
        self.btn_play = QPushButton("Play")
        self.btn_stop = QPushButton("Stop")
        self.btn_reset = QPushButton("Reset")
        self.btn_record = QPushButton("Start Recording")
        self.btn_record.setStyleSheet("background-color: green; color: white;")

        btn_layout.addWidget(self.btn_play)
        btn_layout.addWidget(self.btn_stop)
        btn_layout.addWidget(self.btn_reset)
        btn_layout.addWidget(self.btn_record)
        control_layout.addLayout(btn_layout)

        self.btn_play.clicked.connect(self.send_play)
        self.btn_stop.clicked.connect(self.send_stop)
        self.btn_reset.clicked.connect(self.handle_reset)
        self.btn_record.clicked.connect(self.toggle_recording)

        # Params row
        param_layout = QHBoxLayout()
        param_layout.addWidget(QLabel("Consigne:"))
        self.input_con = QLineEdit("25.0")
        param_layout.addWidget(self.input_con)

        param_layout.addWidget(QLabel("Ki:"))
        self.input_Ki = QLineEdit("1.0")
        param_layout.addWidget(self.input_Ki)

        param_layout.addWidget(QLabel("Kp:"))
        self.input_Kp = QLineEdit("0.5")
        param_layout.addWidget(self.input_Kp)

        self.btn_param = QPushButton("Send Param")
        param_layout.addWidget(self.btn_param)
        self.btn_param.clicked.connect(self.send_param)
        control_layout.addLayout(param_layout)

        # Raw command
        raw_cmd_layout = QHBoxLayout()
        raw_cmd_layout.addWidget(QLabel("Raw Command:"))
        self.input_raw_cmd = QLineEdit("PARAM C=25.0 I=1.0 K=0.5")
        raw_cmd_layout.addWidget(self.input_raw_cmd)
        self.btn_send_raw = QPushButton("Send")
        raw_cmd_layout.addWidget(self.btn_send_raw)
        self.btn_send_raw.clicked.connect(self.send_raw_cmd)
        control_layout.addLayout(raw_cmd_layout)

        # Text area
        self.text_area = QPlainTextEdit()
        self.text_area.setReadOnly(True)
        control_layout.addWidget(self.text_area)

        # **Add a stability label** here (outside the graph)
        self.lbl_stability = QLabel("Stability: unknown")
        self.lbl_stability.setStyleSheet("color: red;")
        control_layout.addWidget(self.lbl_stability)

        # Plots
        self.temp_plot = TempPlotWidget()
        self.cmd_plot  = CommandPlotWidget()

        plot_layout.addWidget(self.temp_plot)
        plot_layout.addWidget(self.cmd_plot)

        main_layout.addLayout(control_layout, 1)
        main_layout.addLayout(plot_layout, 2)
        self.setLayout(main_layout)

    #############################################
    # BACKGROUND DATA HANDLER
    #############################################
    def on_line_received(self, line):
        self.lineReceivedText.emit(f"[RX] {line}")
        self.lineReceivedPlot.emit(line)

        # Check stability each time we see T3
        self.check_stability_zone(line)

        if self.recording:
            self.excel_recorder.parse_and_write(line)

    def append_line(self, text):
        self.text_area.appendPlainText(text)

    #############################################
    # STABILITY LOGIC
    #############################################
    def check_stability_zone(self, line):
        mt3e = re.search(r"t3\s*est:\s*([\d\.]+)", line)
        if not mt3e:
            return
        t3_val = float(mt3e.group(1))

        lower = self.current_setpoint - self.allowable_error
        upper = self.current_setpoint + self.allowable_error


        now = time.time()
        in_zone_now = (lower <= t3_val <= upper)

        if in_zone_now and not self.in_stable_zone:
            # just entered
            self.in_stable_zone = True
            self.enter_time = now
            self.lbl_stability.setStyleSheet("color: green;")

        elif in_zone_now and self.in_stable_zone:
            # still in zone
            elapsed = now - self.enter_time
            self.lbl_stability.setStyleSheet("color: green;")
            if elapsed >= 5 * 60:
                # stable for 5 min
                self.lbl_stability.setText(f"Stability: stable ≥ 5 min!, elapsed time : {elapsed} seconds.")
            else:
                self.lbl_stability.setText(f"Stability: in zone {int(elapsed)}s (need 300s).")
        else:
            # out of zone
            if self.in_stable_zone:
                # we just left
                self.lbl_stability.setText(f"Stability: not stable. Need to be between : {lower} and {upper}.")
            self.in_stable_zone = False
            self.lbl_stability.setStyleSheet("color: red;")
            self.enter_time = None

    def get_current_t3_est(self):
        """Return the last T3 from temp_plot or fallback."""
        if len(self.temp_plot.t3_est_values) > 0:
            return self.temp_plot.t3_est_values[-1]
        return 25.0

    #############################################
    # UI Commands
    #############################################
    def send_play(self):
        self._send_line("p")

    def send_stop(self):
        self._send_line("S")

    def handle_reset(self):
        self.send_reset()
        self.temp_plot.reset_plot()
        self.cmd_plot.reset_plot()
        self.lbl_stability.setText("Stability: reset.")

    def send_reset(self):
        self._send_line("PARAM C=25.0 I=1.0 K=0.5")

    def send_param(self):
        # 1) read new setpoint, recalc error
        con = float(self.input_con.text().strip())
        old_temp = self.get_current_t3_est()
        diff = abs(old_temp - con)
        self.allowable_error = 0.05 * diff
        
        if self.allowable_error <= 0.1:
            self.allowable_error = 0.1

        self.current_setpoint = con

        # 2) reset timer
        self.enter_time = None
        self.in_stable_zone = False
        self.lbl_stability.setText(f"Stability: changed setpoint. initial:{old_temp}, allowable:{self.allowable_error} ")

        # 3) send param
        Ki = self.input_Ki.text().strip()
        Kp = self.input_Kp.text().strip()
        cmd = f"PARAM C={con} I={Ki} K={Kp}"
        self._send_line(cmd)
        self.excel_recorder.set_consigne(con)

    def send_raw_cmd(self):
        raw_cmd = self.input_raw_cmd.text().strip()
        self._send_line(raw_cmd)

    def _send_line(self, data_str):
        if self.ser and self.ser.is_open:
            line = data_str + "\n"
            self.ser.write(line.encode('ascii', errors='ignore'))
            self.text_area.appendPlainText(f"[TX] {data_str}")

    #############################################
    # RECORDING
    #############################################
    def toggle_recording(self):
        if self.recording:
            self.recording = False
            self.excel_recorder.save_and_close()
            self.text_area.appendPlainText("Stopped recording to Excel.")
            self.btn_record.setText("Start Recording")
            self.btn_record.setStyleSheet("background-color: green; color: white;")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_xlsx = os.path.join(self.data_dir, f"Xlsx/data_{timestamp}.xlsx")
            self.excel_recorder.create_copy_and_open(new_xlsx)
            self.excel_recorder.set_consigne(self.current_setpoint)
            self.recording = True
            self.text_area.appendPlainText(f"Recording to Excel: {new_xlsx}")
            self.btn_record.setText("Recording")
            self.btn_record.setStyleSheet("background-color: red; color: white;")

    #############################################
    # Cleanup
    #############################################
    def closeEvent(self, event):
        if hasattr(self, 'read_thread') and self.read_thread:
            self.read_thread.stop()
            self.read_thread.join()
        if hasattr(self, 'ser') and self.ser and self.ser.is_open:
            self.ser.close()

        self.excel_recorder.save_and_close()
        event.accept()


def main():
    app = QApplication(sys.argv)
    window = SerialMonitor(port="COM9", baudrate=115200)

    screen = app.primaryScreen().availableGeometry()
    w = int(screen.width() * 0.8)
    h = int(screen.height() * 0.8)
    x = screen.x() + (screen.width() - w)//2
    y = screen.y() + (screen.height() - h)//2
    window.setGeometry(x, y, w, h)

    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
