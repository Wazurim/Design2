import re
import shutil
from openpyxl import load_workbook

class ExcelRecorder:
    """
    Handles copying a template workbook and writing rows of data
    on sheet #2 ("Réponse échelon").
    """

    # Regex patterns to extract info from each serial line:
    regex_time_s  = re.compile(r"([\d\.]+)\s*s")   # e.g. "4.0 s"
    regex_time_ms = re.compile(r"([\d\.]+)\s*ms")  # e.g. "4000.0 ms"

    regex_pwm = re.compile(r"PWM duty:\s*([\d]+)\s*/\s*([\d]+)")
    regex_t1  = re.compile(r"ADC t1:\s*([\d\.]+)")
    regex_t2  = re.compile(r"ADC t2:\s*([\d\.]+)")
    regex_t3  = re.compile(r"ADC t3:\s*([\d\.]+)")
    regex_t3_est = re.compile(r"ADC t3 estimate:\s*([\d\.]+)")
    regex_t4  = re.compile(r"ADC t4:\s*([\d\.]+)")

    def __init__(self, template_path, sheet_name="data"):
        self.template_path = template_path
        self.current_path = None
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.row_index = 2      # Start writing at row 2 (row 1 is typically headers)
        self.base_time = None   # For offsetting time so first data = 0
        self.t3_values = []     # For computing T3 moyen
        self.consigne = 25.0    # Example default setpoint

    def create_copy_and_open(self, new_path):
        """
        Copy the template to 'new_path', then open it for writing.
        """
        shutil.copyfile(self.template_path, new_path)
        self.workbook = load_workbook(new_path)
        self.current_path = new_path
        self.sheet = self.workbook[self.sheet_name]
        self.row_index = 2
        self.base_time = None
        self.t3_values.clear()

    def set_consigne(self, value):
        """Let the main app update the consigne if needed."""
        self.consigne = float(value)

    def parse_and_write(self, line):
        """
        Parse the given serial line for time, PWM, T1..T4, optional T3 estimate, etc.
        Then write a row into the Excel sheet if T1..T4 are found.
        Order: Time, Consigne, U, T1, T2, T3, T3_estime, T4, T3_moyen
        """

        # 1) Parse time in seconds or ms:
        time_seconds = None
        match_s  = self.regex_time_s.search(line)
        match_ms = self.regex_time_ms.search(line)

        if match_s:
            time_seconds = float(match_s.group(1))  # e.g. "4.0 s"
        elif match_ms:
            ms_val = float(match_ms.group(1))       # e.g. "4000.0 ms"
            time_seconds = ms_val / 1000.0

        if time_seconds is None:
            return  # No time => skip

        # 2) PWM duty => needed for U
        pwm_match = self.regex_pwm.search(line)
        if not pwm_match:
            return
        duty_val = float(pwm_match.group(1))
        duty_max = float(pwm_match.group(2))
        fraction = duty_val / duty_max if duty_max != 0 else 0.0
        U = fraction * 10.0 - 5.0  # => -5..+5 V range

        # 3) T1..T4 (required) + T3 estimate (optional)
        m_t1  = self.regex_t1.search(line)
        m_t2  = self.regex_t2.search(line)
        m_t3  = self.regex_t3.search(line)
        m_t4  = self.regex_t4.search(line)

        if not (m_t1 and m_t2 and m_t3 and m_t4):
            return  # Must have at least T1..T4

        # If T3 estimate is missing, we'll set it to blank in Excel
        m_t3e = self.regex_t3_est.search(line)
        t3_est_val = float(m_t3e.group(1)) if m_t3e else None

        t1_val = float(m_t1.group(1))
        t2_val = float(m_t2.group(1))
        t3_val = float(m_t3.group(1))
        t4_val = float(m_t4.group(1))

        # 4) If this is the first line => store base_time
        if self.base_time is None:
            self.base_time = time_seconds

        temps = time_seconds - self.base_time

        # 5) T3 moyen => average of all T3 so far
        self.t3_values.append(t3_val)
        if  len(self.t3_values) > 20:
            t3_moy = sum(self.t3_values[-20:]) / 20
        else:
            t3_moy = sum(self.t3_values) / len(self.t3_values)

        # 6) Write row [Time, Consigne, U, T1, T2, T3, T3_est, T4, T3_moy]
        self.sheet.cell(row=self.row_index, column=1).value = temps
        self.sheet.cell(row=self.row_index, column=2).value = self.consigne
        self.sheet.cell(row=self.row_index, column=3).value = U
        self.sheet.cell(row=self.row_index, column=4).value = t1_val
        self.sheet.cell(row=self.row_index, column=5).value = t2_val
        self.sheet.cell(row=self.row_index, column=6).value = t3_val
        # T3 estimate is optional, set to "" if missing
        self.sheet.cell(row=self.row_index, column=7).value = t3_est_val if t3_est_val else ""
        self.sheet.cell(row=self.row_index, column=8).value = t4_val
        self.sheet.cell(row=self.row_index, column=9).value = t3_moy

        self.row_index += 1

    def save_and_close(self):
        """Save the workbook and clear references."""
        if self.workbook:
            self.workbook.save(self.current_path)
            self.workbook = None
            self.sheet = None
