import re
import sys
import os
import time
import serial
import threading
from datetime import datetime

import matplotlib
matplotlib.use("Qt5Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

from PyQt5.QtCore import Qt, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QPlainTextEdit
)

################################################################################
# ExcelRecorder (unchanged from before, adapted for your new line format)
################################################################################
class ExcelRecorder:
    """ Class to record data in an Excel file.
    This class is responsible for creating a copy of the template file and writing data to it.
    It uses openpyxl to handle Excel files.

    """
    regex_time_s  = re.compile(r"([\d\.]+)\s*s")   
    regex_pwm     = re.compile(r"PWM\s*:\s*([\d]+)\s*/\s*([\d]+)")
    regex_t1      = re.compile(r"t1:\s*([\d\.]+)")
    regex_t2      = re.compile(r"t2:\s*([\d\.]+)")
    regex_t3      = re.compile(r"t3:\s*([\d\.]+)")
    regex_t3_est  = re.compile(r"t3\s*est:\s*([\d\.]+)")
    regex_t4      = re.compile(r"t4:\s*([\d\.]+)")

    def __init__(self, template_path, sheet_name="data"):
        """ Initialize the ExcelRecorder with the template path and sheet name.

        Args:
            template_path (String): Path to the template Excel file.
            sheet_name (str, optional): Name of the sheet to write data to. Defaults to "data".
        """
        self.template_path = template_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.current_path = None
        self.row_index = 2
        self.base_time = None
        self.t3_values = []
        self.consigne = 25.0

    def create_copy_and_open(self, new_path):
        """ Create a copy of the template file and open it for writing.
        This method will copy the template file to a new path and open it using openpyxl.
        It will also reset the row index and clear the t3 values.

        Args:
            new_path (String): Path to the new Excel file to create.
        """
        import shutil
        from openpyxl import load_workbook

        shutil.copyfile(self.template_path, new_path)
        self.current_path = new_path

        self.workbook = load_workbook(new_path)
        self.sheet = self.workbook[self.sheet_name]
        self.row_index = 2
        self.base_time = None
        self.t3_values.clear()

    def set_consigne(self, value):
        """ Set the consigne value for the Excel file.

        Args:
            value (float): Consigne value to set.
        """
        self.consigne = float(value)

    def parse_and_write(self, line):
        """ Parse the line received from the serial port and write data to the Excel file.
        This method will extract the relevant data from the line and write it to the Excel file.
        It will also calculate the time difference from the base time and update the row index.


        Args:
            line (_type_): Line received from the serial port.
        """
        m_s = self.regex_time_s.search(line)
        print(line)
        if not m_s:
            return
        time_s = float(m_s.group(1))

        pwm_m = self.regex_pwm.search(line)
        if not pwm_m:
            return
        duty_val = float(pwm_m.group(1))
        duty_max = float(pwm_m.group(2))
        fraction = duty_val / duty_max if duty_max else 0.0
        U = fraction * 10.0 - 5.0

        mt1 = self.regex_t1.search(line)
        mt2 = self.regex_t2.search(line)
        mt3 = self.regex_t3.search(line)
        mt4 = self.regex_t4.search(line)
        if not (mt1 and mt2 and mt3 and mt4):
            return

        t1_val = float(mt1.group(1))
        t2_val = float(mt2.group(1))
        t3_val = float(mt3.group(1))
        t4_val = float(mt4.group(1))

        mt3e = self.regex_t3_est.search(line)
        t3_est_val = float(mt3e.group(1)) if mt3e else None

        if self.base_time is None:
            self.base_time = time_s

        temps = time_s - self.base_time
        self.t3_values.append(t3_val)
        t3_moy = sum(self.t3_values) / len(self.t3_values)

        # Write row
        row = self.row_index
        self.sheet.cell(row=row, column=1).value = temps
        self.sheet.cell(row=row, column=2).value = self.consigne
        self.sheet.cell(row=row, column=3).value = U
        self.sheet.cell(row=row, column=4).value = t1_val
        self.sheet.cell(row=row, column=5).value = t2_val
        self.sheet.cell(row=row, column=6).value = t3_val
        self.sheet.cell(row=row, column=7).value = t3_est_val if t3_est_val else ""
        self.sheet.cell(row=row, column=8).value = t4_val
        self.sheet.cell(row=row, column=9).value = t3_moy

        self.row_index += 1

    def save_and_close(self):
        """ Save the Excel file and close it.
        This method will save the current workbook and close it. It will also reset the workbook and sheet attributes.

        """
        if self.workbook:
            self.workbook.save(self.current_path)
            self.workbook = None
            self.sheet = None


################################################################################
# Thread for reading data from the serial port
################################################################################
class SerialReadThread(threading.Thread):
    """ Thread to read data from the serial port.
    This thread will continuously read lines from the serial port and emit a signal with the line data.

    Args:
        threading (Thread): Thread class from the threading module.
    """
    def __init__(self, ser, callback):
        """ Initialize the SerialReadThread with the serial port and callback function.
        This method will set the serial port and callback function, and initialize the running flag.


        Args:
            ser (serial.Serial):  serial.Serial object: Serial port to read data from.
            callback (function): Callback function to call with the line data.
        """
        super().__init__()
        self.ser = ser
        self.callback = callback
        self.running = True

    def run(self):
        """ Run the thread to read data from the serial port.
        """
        while self.running:
            try:
                line = self.ser.readline().decode('ascii', errors='ignore').strip()
                if line:
                    self.callback(line)
            except Exception as e:
                print("Serial error:", e)
                break
        print("Serial reading thread terminated.")

    def stop(self):
        """ Stop the thread and close the serial port.
        """
        self.running = False


################################################################################
# Live Plot #1: Temperatures (T1..T4 + T3_est) [no stability text here]
################################################################################
class TempPlotWidget(QWidget):
    """ Widget to plot the temperatures in real-time.

    Args:
        QWidget (QWidget): QWidget class from the PyQt5 module.
    """
    def __init__(self, parent=None):
        """
        Initialize the TempPlotWidget with the parent widget.

        Args:
            parent (_type_, optional): Parent widget. Defaults to None.
        """
        super().__init__(parent)

        # Data arrays
        self.time_values     = []
        self.consigne_values = []
        self.t1_values       = []
        self.t2_values       = []
        self.t3_values       = []
        self.t4_values       = []
        self.t3_est_values   = []


        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)

        self.line_consigne,     = self.ax.plot([], [], label="Consigne")
        self.line_t1,     = self.ax.plot([], [], label="T1")
        self.line_t2,     = self.ax.plot([], [], label="T2")
        self.line_t3,     = self.ax.plot([], [], label="T3")
        self.line_t4,     = self.ax.plot([], [], label="T4")
        self.line_t3_est, = self.ax.plot([], [], label="T3_est")

        self.ax.set_xlabel("Time (s)")
        self.ax.set_ylabel("Temperature (°C)")
        self.ax.set_title("Real-Time Temps + T3_est")
        self.ax.grid(True)
        self.ax.legend()

        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.setLayout(layout)

    def get_moyt3_est(self):
        """ Get the average of the last 30 T3_est values.

        Returns:
            float:  Average of the last 30 T3_est values or 0 if no values are available.
        """
        
        if  len(self.t3_values) > 30:
            t3_moy = sum(self.t3_values[-30:]) / 30

        elif len(self.t3_values) == 0:
            t3_moy = 0
        else:
            t3_moy = sum(self.t3_values) / len(self.t3_values)

        return t3_moy


    @pyqtSlot(str)
    def parse_and_update(self, line):
        """ Parse the line received from the serial port and update the plot.
        This method will extract the relevant data from the line and update the plot with the new data.
        It will also adjust the y-axis limits based on the data.


        Args:
            line (_type_):  Line received from the serial port.
        """
        match_s = re.search(r"([\d\.]+)\s*s", line)
        if not match_s:
            return
        time_s = float(match_s.group(1))
   

        # t1..t4
        mt1 = re.search(r"t1:\s*([\d\.]+)", line)
        mt2 = re.search(r"t2:\s*([\d\.]+)", line)
        mt3 = re.search(r"t3:\s*([\d\.]+)", line)
        mt4 = re.search(r"t4:\s*([\d\.]+)", line)
        if not (mt1 and mt2 and mt3 and mt4):
            return
        
        t1_val = float(mt1.group(1))
        t2_val = float(mt2.group(1))
        t3_val = float(mt3.group(1))
        t4_val = float(mt4.group(1))

        # t3 est (optional)
        mt3e = re.search(r"t3\s*est:\s*([\d\.]+)", line)
        t3_est_val = float(mt3e.group(1)) if mt3e else None

        consigne = re.search(r"consigne:\s*([\d\.]+)", line)
        consigne_val = float(consigne.group(1)) if consigne else None
        
        # Append
        self.time_values.append(time_s)
        self.t1_values.append(t1_val)
        self.t2_values.append(t2_val)
        self.t3_values.append(t3_val)
        self.t4_values.append(t4_val)

        if t3_est_val is not None:
            self.t3_est_values.append(t3_est_val)
        else:
            self.t3_est_values.append(None)
            
        if consigne_val is not None:
            self.consigne_values.append(consigne_val)
        else:
            self.consigne_values.append(None)

        all_y = self.t1_values + self.t2_values + self.t3_est_values + self.t3_values + self.t4_values + self.consigne_values
        all_y = [x for x in all_y if x is not None]
        y_min, y_max = min(all_y), max(all_y)
        self.ax.set_ylim(y_min-1, y_max+1)

        # Update lines
        self.line_consigne.set_xdata(self.time_values)
        self.line_consigne.set_ydata([val if val is not None else float("nan") for val in self.consigne_values])
        self.line_t1.set_xdata(self.time_values)
        self.line_t1.set_ydata(self.t1_values)
        self.line_t2.set_xdata(self.time_values)
        self.line_t2.set_ydata(self.t2_values)
        self.line_t3.set_xdata(self.time_values)
        self.line_t3.set_ydata(self.t3_values)

        # T3_est -> None => NaN
        t3_est_y = [val if val is not None else float("nan") for val in self.t3_est_values]
        self.line_t3_est.set_xdata(self.time_values)
        self.line_t3_est.set_ydata(t3_est_y)

        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()

    def reset_plot(self):
        """ Reset the plot data and clear the lines.
        """
        self.time_values.clear()
        self.consigne_values.clear()
        self.t1_values.clear()
        self.t2_values.clear()
        self.t3_values.clear()
        self.t4_values.clear()
        self.t3_est_values.clear()

        self.line_consigne.set_data([], [])
        self.line_t1.set_data([], [])
        self.line_t2.set_data([], [])
        self.line_t3.set_data([], [])
        self.line_t4.set_data([], [])
        self.line_t3_est.set_data([], [])

        self.ax.relim()

        self.canvas.draw()


################################################################################
# Live Plot #2: Command (U)
################################################################################
class CommandPlotWidget(QWidget):
    """ Widget to plot the command U in real-time.


    Args:
        QWidget (QWidget): QWidget class from the PyQt5 module.
    """
    def __init__(self, parent=None):
        """_summary_

        Args:
            parent (QWidget, optional): Parent widget. Defaults to None.
        """
        super().__init__(parent)
        self.time_values = []
        self.u_values    = []

        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)

        self.line_u, = self.ax.plot([], [], label="U (%)")

        self.ax.set_xlabel("Time (s)")
        self.ax.set_ylabel("Command U (%)")
        self.ax.set_title("Real-Time Command U")
        self.ax.grid(True)
        self.ax.legend()

        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.setLayout(layout)

    @pyqtSlot(str)
    def parse_and_update(self, line):
        """ Parse the line received from the serial port and update the plot.
        This method will extract the relevant data from the line and update the plot with the new data.
        It will also adjust the y-axis limits based on the data.


        Args:
            line (_type_):  Line received from the serial port.
        """
        match_s = re.search(r"([\d\.]+)\s*s", line)
        if not match_s:
            return
        time_s = float(match_s.group(1))


        pwm_m = re.search(r"PWM\s*:\s*([\d]+)\s*/\s*([\d]+)", line)
        if not pwm_m:
            return
        duty_val = float(pwm_m.group(1))
        duty_max = float(pwm_m.group(2))
        ratio = duty_val/duty_max if duty_max else 0.0
        u_val = -ratio * 200 + 100

        self.time_values.append(time_s)
        self.u_values.append(u_val)

        self.line_u.set_xdata(self.time_values)
        self.line_u.set_ydata(self.u_values)

        self.ax.autoscale_view()

        self.ax.relim()

        self.canvas.draw()

    def reset_plot(self):
        """
        Reset the plot data and clear the lines.
        """
        self.time_values.clear()
        self.u_values.clear()

        self.line_u.set_data([], [])
        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()


################################################################################
# Main PyQt Window
################################################################################
class SerialMonitor(QWidget):
    """ Class for the main window of the serial monitor application.
    This class is responsible for creating the main window and its layout.
    It contains all the input fields for the simulation parameters and the buttons to start the simulation and save the parameters.


    Args:
        QWidget (QWidget):  QWidget class from the PyQt5 module.

    Raises:
        RuntimeError: Could not open port.

    """
    lineReceivedText = pyqtSignal(str)
    lineReceivedPlot = pyqtSignal(str)

    def __init__(self, port="COM4", baudrate=115200, parent=None):
        """ Initialize the SerialMonitor with the given port and baudrate.
        This method will set the port and baudrate, and initialize the serial port.
        It will also create the UI elements and connect the signals to the slots.
        It will also create the ExcelRecorder object to record data in an Excel file.
        It will also create the stability tracking variables and connect the signals to the slots.

        Args:
            port (str, optional): Port to open. Defaults to "COM4".
            baudrate (int, optional): Baudrate for the serial port. Defaults to 115200.
            parent (_type_, optional): Parent widget. Defaults to None.

        Raises:
            RuntimeError: Could not open port.
        """
        super().__init__(parent)
        self.setWindowTitle("Design 2 Prototype serial monitor")
        self.setWindowIcon(QIcon("icon.png"))

        self.data_dir = os.path.join(os.getcwd(), "data")
        os.makedirs(self.data_dir, exist_ok=True)

        self.recording = False
        self.file = None

        self.template_file = os.path.join(os.getcwd(), "app/assets/data.xlsx")
        self.excel_recorder = ExcelRecorder(self.template_file)

        # For stability tracking
        self.current_setpoint = 25.0
        self.allowable_error  = 0.8
        self.in_stable_zone   = False
        self.in_precise_zone   = False
        self.enter_time       = None
        self.time_counting = False

        self._build_ui()

        self.lineReceivedText.connect(self.append_line)
        self.lineReceivedPlot.connect(self.temp_plot.parse_and_update)
        self.lineReceivedPlot.connect(self.cmd_plot.parse_and_update)

        # Try open serial
        try:
            self.ser = serial.Serial(port, baudrate, timeout=1)
            time.sleep(2)
        except Exception as e:
            raise RuntimeError(f"Could not open port {port}: {e}")

        self.read_thread = SerialReadThread(self.ser, self.on_line_received)
        self.read_thread.start()

    def _build_ui(self):
        """ Build the UI for the main window.
        This method will create the main layout and add all the UI elements to it.

        """
        main_layout = QHBoxLayout()
        control_layout = QVBoxLayout()
        plot_layout = QVBoxLayout()

        # Buttons
        btn_layout = QHBoxLayout()
        self.btn_play = QPushButton("Play")
        self.btn_play.setStyleSheet("background-color: white; color: black;")
        self.btn_play.setText("Send and play")
        self.btn_stop = QPushButton("Stop")
        self.btn_reset = QPushButton("Reset")
        self.btn_record = QPushButton("Start Recording")
        self.btn_record.setStyleSheet("background-color: green; color: white;")

        btn_layout.addWidget(self.btn_play)
        btn_layout.addWidget(self.btn_stop)
        btn_layout.addWidget(self.btn_reset)
        btn_layout.addWidget(self.btn_record)
        control_layout.addLayout(btn_layout)

        self.btn_play.clicked.connect(self.send_play)
        self.btn_stop.clicked.connect(self.send_stop)
        self.btn_reset.clicked.connect(self.handle_reset)
        self.btn_record.clicked.connect(self.toggle_recording)

        # Params row
        param_layout = QHBoxLayout()
        param_layout.addWidget(QLabel("Consigne:"))
        self.input_con = QLineEdit("25.0")
        param_layout.addWidget(self.input_con)

        param_layout.addWidget(QLabel("P:"))
        self.input_p = QLineEdit("4.9931")
        param_layout.addWidget(self.input_p)

        param_layout.addWidget(QLabel("I:"))
        self.input_i = QLineEdit("0.008689")
        param_layout.addWidget(self.input_i)

        param_layout.addWidget(QLabel("D:"))
        self.input_d = QLineEdit("48.2037")
        param_layout.addWidget(self.input_d)

        param_layout.addWidget(QLabel("F:"))
        self.input_f = QLineEdit("4.964471")
        param_layout.addWidget(self.input_f)
        control_layout.addLayout(param_layout)


        self.lbl_precision = QLabel("Precision: unknown")
        self.lbl_precision.setStyleSheet("color: red;")
        control_layout.addWidget(self.lbl_precision)

        self.lbl_stabilite = QLabel("Stability: unknown")
        self.lbl_stabilite.setStyleSheet("color: red;")
        control_layout.addWidget(self.lbl_stabilite)

        self.lbl_timer = QLabel("N/A")
        self.lbl_timer.setStyleSheet("color: red;")
        control_layout.addWidget(self.lbl_timer)

        # Text area
        self.text_area = QPlainTextEdit()
        self.text_area.setReadOnly(True)
        control_layout.addWidget(self.text_area)

        # Plots
        self.temp_plot = TempPlotWidget()
        self.cmd_plot  = CommandPlotWidget()

        plot_layout.addWidget(self.temp_plot)
        plot_layout.addWidget(self.cmd_plot)

        main_layout.addLayout(control_layout, 1)
        main_layout.addLayout(plot_layout, 2)
        self.setLayout(main_layout)

    #############################################
    # BACKGROUND DATA HANDLER
    #############################################
    def on_line_received(self, line):
        """ Handle the line received from the serial port.
        This method will emit a signal with the line data and parse it to update the plots.

        Args:
            line (_type_): Line received from the serial port.
        """
        self.lineReceivedText.emit(f"[RX] {line}")
        self.lineReceivedPlot.emit(line)

        # Check stability each time we see T3
        self.check_stability_zone(line)

        if self.recording:
            self.excel_recorder.parse_and_write(line)

    def append_line(self, text):
        """ Append a line to the text area.
        This method will append the given text to the text area in the UI.

        Args:
            text (string): Text to append.
        """
        self.text_area.appendPlainText(text)

    #############################################
    # STABILITY LOGIC
    #############################################
    def check_stability_zone(self, line):
        """ Check if the T3 value is within the stability zone.
        This method will parse the line to extract the T3 value and check if it is within the stability zone.
        It will also update the labels in the UI to indicate the stability status.

        Args:
            line (_type_):  Line received from the serial port.
        """
        
        mt3e = re.search(r"t3\s*est:\s*([\d\.]+)", line)
        if not mt3e:
            return
        t3_val = float(mt3e.group(1))

        lower = self.current_setpoint - self.allowable_error
        upper = self.current_setpoint + self.allowable_error


        now = time.time()
        in_zone_now = (lower <= t3_val and t3_val <= upper)

        if in_zone_now:
            # just entered
            self.in_precise_zone = True
            self.lbl_precision.setStyleSheet("color: green;")
            self.lbl_precision.setText(f"Precision: in zone.")

        else:
            self.lbl_precision.setText(f"Precision: Bad, need to be between : {lower} and {upper}.")
            self.in_precise_zone = False
            self.time_counting = False
            self.lbl_precision.setStyleSheet("color: red;")



        if len(self.temp_plot.t3_est_values) > 20:
            last_values_t3_est = self.temp_plot.t3_est_values[-20:]
            diff = max(last_values_t3_est) -min(last_values_t3_est)

            if diff < 0.1:
                self.lbl_stabilite.setText(f"Stabilite: Stable, ecart < 0.1 depuis plus de 20 secondes :)")
                self.lbl_stabilite.setStyleSheet("color: green;")
                self.in_stable_zone = True

            else:
                self.lbl_stabilite.setText(f"Stabilite: instable :(   Ecart: {diff}")   
                self.lbl_stabilite.setStyleSheet("color: red;")
                self.in_stable_zone = False
                self.time_counting = False
                
        else:
            self.lbl_stabilite.setText(f"Stabilite: calibrating")   
            self.lbl_stabilite.setStyleSheet("color: orange;")
            self.in_stable_zone = False
            self.time_counting = False
            

        if self.in_stable_zone and self.in_precise_zone and not self.time_counting:
            self.enter_time = now - 20
            self.lbl_timer.setStyleSheet("color: green;")
            self.time_counting = True
            
        elif self.in_stable_zone and self.in_precise_zone and self.time_counting:
            elapsed = now - self.enter_time
            self.lbl_timer.setText(f"In zones :) : elapsed: {elapsed}")
        else:
            self.enter_time = now + 20
            self.lbl_timer.setStyleSheet("color: red;")
            self.lbl_timer.setText(f"N/A")

                    

    def get_current_t3_est(self):
        """ Get the current T3_est value.
        This method will return the last T3_est value from the plot data.

        Returns:
            float:  Current T3_est value or 25.0 if no values are available.
        """
        if len(self.temp_plot.t3_est_values) > 0:
            return self.temp_plot.t3_est_values[-1]
        return 25.0

    #############################################
    # UI Commands
    #############################################
    def send_play(self):
        """ Send the play command to the serial port.
        """
        self._send_line("p")
        self.send_param()
        self.btn_play.setStyleSheet("background-color: lightblue; color: white;")
        self.btn_play.setText("Update param")

    def send_stop(self):
        """ Send the stop command to the serial port.
        """
        self._send_line("S")
        self.btn_play.setStyleSheet("background-color: white; color: black;")
        self.btn_play.setText("Send and play")

    def handle_reset(self):
        """ Handle the reset command.
        This method will reset the plot data and send the reset command to the serial port.
        """
        self.send_reset()
        self.temp_plot.reset_plot()
        self.cmd_plot.reset_plot()
        self.lbl_precision.setText("Stability: reset.")

    def send_reset(self):
        """ Send the reset command to the serial port.
        This method will send the reset command to the serial port and reset the plot data.
        """
        self._send_line("PARAM C=25.0 I=1.0 K=0.5")

    def send_param(self):
        """ Send the parameters to the serial port.
        This method will read the parameters from the input fields and send them to the serial port.
        """
        # 1) read new setpoint, recalc error
        con = float(self.input_con.text().strip())
        old_temp = self.get_current_t3_est()
        self.allowable_error = 0.8


        self.current_setpoint = con

        # 2) reset timer
        self.enter_time = None
        self.in_stable_zone = False
        self.lbl_precision.setText(f"Precision: changed setpoint. initial:{old_temp}, allowable:{self.allowable_error} ")

        # 3) send param
        i = self.input_i.text().strip()
        p = self.input_p.text().strip()
        d = self.input_d.text().strip()
        f = self.input_f.text().strip()

        cmd = f"PARAM C={con} P={p} I={i} D={d} F={f}"
        self._send_line(cmd)
        self.excel_recorder.set_consigne(con)

    def _send_line(self, data_str):
        """ Send a line to the serial port.
        This method will encode the line and send it to the serial port.

        Args:
            data_str (string):  Line to send.
        """
        if self.ser and self.ser.is_open:
            line = data_str + "\n"
            self.ser.write(line.encode('ascii', errors='ignore'))
            self.text_area.appendPlainText(f"[TX] {data_str}")

    #############################################
    # RECORDING
    #############################################
    def toggle_recording(self):
        """ Toggle the recording state.
        This method will start or stop the recording to the Excel file.
        """
        if self.recording:
            self.recording = False
            self.excel_recorder.save_and_close()
            self.text_area.appendPlainText("Stopped recording to Excel.")
            self.btn_record.setText("Start Recording")
            self.btn_record.setStyleSheet("background-color: green; color: white;")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_xlsx = os.path.join(self.data_dir, f"Xlsx/data_{timestamp}.xlsx")
            self.excel_recorder.create_copy_and_open(new_xlsx)
            self.excel_recorder.set_consigne(self.current_setpoint)
            self.recording = True
            self.text_area.appendPlainText(f"Recording to Excel: {new_xlsx}")
            self.btn_record.setText("Recording")
            self.btn_record.setStyleSheet("background-color: red; color: white;")

    #############################################
    # Cleanup
    #############################################
    def closeEvent(self, event):
        """ Handle the close event for the main window.
        This method will stop the serial reading thread and close the serial port.

        Args:
            event (event): event object for the close event.
        """
        if hasattr(self, 'read_thread') and self.read_thread:
            self.read_thread.stop()
            self.read_thread.join()
        if hasattr(self, 'ser') and self.ser and self.ser.is_open:
            self.ser.close()

        self.excel_recorder.save_and_close()
        event.accept()


def main():
    app = QApplication(sys.argv) #create the application
    window = SerialMonitor(port="COM9", baudrate=115200) #create the main window

    screen = app.primaryScreen().availableGeometry() # Get the screen geometry
    w = int(screen.width() * 0.95) # Set the width to 95% of the screen width
    h = int(screen.height() * 0.95) # Set the height to 95% of the screen height
    x = screen.x() + (screen.width() - w)//2 # Center the window horizontally
    y = screen.y() + (screen.height() - h)//2 # Center the window vertically
    window.setGeometry(x, y, w, h) # Set the window geometry

    window.show() #show the main window
    sys.exit(app.exec_()) #run the application


if __name__ == "__main__":
    main()  
